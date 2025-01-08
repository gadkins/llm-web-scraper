import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from pydantic import BaseModel
from typing import Optional, Dict, Union, List

# Help functions for saving Pydantic model to CSV

def parse_model_to_filtered_dict(model_instance: BaseModel, field_mapping: Dict[str, str]) -> Dict[str, Optional[str]]:
    """
    Parse a Pydantic model instance into a dictionary with fields filtered and mapped to desired column names.
    """
    # Reverse the field mapping to go from model field -> display name
    reverse_mapping = {v: k for k, v in field_mapping.items()}
    
    # Get the model data
    model_data = model_instance.model_dump()
    
    # Create the filtered dictionary with display names as keys
    filtered_data = {}
    for field_name, value in model_data.items():
        if field_name in reverse_mapping:
            display_name = reverse_mapping[field_name]
            if isinstance(value, list):
                value = ", ".join(map(str, value))
            filtered_data[display_name] = value
            
    return filtered_data

def save_to_spreadsheet(
    data: Union[Dict, List[Dict], BaseModel, List[BaseModel]],
    field_mapping: Dict[str, Optional[str]],
    excel_filename: str = "output.xlsx",
    csv_filename: str = "output.csv"
):
    """
    Saves data from one or more instances to Excel and CSV files.
    """
    # Convert input to list if single instance
    if not isinstance(data, list):
        data = [data]
    
    # Get the column headers (display names)
    headers = list(field_mapping.keys())
    
    # Create DataFrame directly from the data
    df = pd.DataFrame(data)
    
    # Ensure all headers exist in DataFrame
    for header in headers:
        if header not in df.columns:
            df[header] = None
            
    # Reorder columns to match headers
    df = df[headers]
    
    # Save to CSV
    df.to_csv(csv_filename, index=False)

    # Save to Excel
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws.append(headers)

    # Write data rows
    for row_dict in data:
        # Create a row with values in the same order as headers
        row_values = []
        for header in headers:
            value = row_dict.get(header)
            if isinstance(value, list):
                value = ", ".join(map(str, value))
            row_values.append(value)
        ws.append(row_values)

    # Apply bold formatting to headers
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    wb.save(excel_filename)

# Append rows to excel
def add_rows_to_excel(
    file_name: str,
    new_data: list,
    field_mapping: Dict[str, Optional[str]]
):
    """
    Adds new rows to an existing Excel file.
    """
    # Convert the list of models to list of dicts if needed
    if isinstance(new_data[0], BaseModel):
        new_data = [
            parse_model_to_filtered_dict(model, field_mapping)
            for model in new_data
        ]

    # Load existing workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Append new rows
    for data in new_data:
        row = []
        for excel_col in field_mapping.keys():  # Use excel column names for consistent ordering
            value = data.get(excel_col)
            if isinstance(value, list):
                value = ", ".join(map(str, value))
            row.append(value)
        ws.append(row)

    wb.save(file_name)
    print(f"New rows successfully added to {file_name}.")